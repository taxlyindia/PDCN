"""
Excel Export Reports — role-based data and remarks visibility.

Role rules:
  FINANCE_TEAM / FINANCE_CFA_TEAM  → All reports, all remarks, sales register, template
  CN_TEAM      → Claims at/after CN stage only, only CN remarks visible
  CFA_TEAM     → Claims at/after finance_approved only, only CFA remarks visible
  DEALER       → Own claims only, all remarks visible (transparency for dealer)
  TENANT_ADMIN → All reports (same as finance)
"""
import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.auth.dependencies import get_active_tenant_user
from app.database import get_db
from app.models import (
    User, UserRole, PDCNClaim, PDCNLineItem,
    SalesRegister, QtyUtilizationLedger, PDCNStatus
)
from app.api.v1.endpoints.pdcn import can_view_claim, get_visible_claims

router = APIRouter(prefix="/reports", tags=["Reports"])


# ─── Style helpers ─────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1E1E3F")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="F8F8FF")
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                      top=BORDER_SIDE, bottom=BORDER_SIDE)

STATUS_COLORS = {
    "draft": "9CA3AF", "submitted": "6366F1", "cn_approved": "06B6D4",
    "cn_rejected": "EF4444", "finance_approved": "10B981",
    "finance_rejected": "F59E0B", "sent_back": "F97316",
    "cn_generated": "0EA5E9", "completed": "8B5CF6", "cancelled": "6B7280",
}

# Stages visible per role
CN_VISIBLE_STATUSES = [
    PDCNStatus.SUBMITTED, PDCNStatus.CN_APPROVED, PDCNStatus.CN_REJECTED,
    PDCNStatus.SENT_BACK, PDCNStatus.FINANCE_APPROVED, PDCNStatus.FINANCE_REJECTED,
    PDCNStatus.CN_PENDING, PDCNStatus.CN_GENERATED, PDCNStatus.COMPLETED,
]
CFA_VISIBLE_STATUSES = [
    PDCNStatus.FINANCE_APPROVED, PDCNStatus.CN_PENDING,
    PDCNStatus.CN_GENERATED, PDCNStatus.COMPLETED,
]


def is_finance(role: str) -> bool:
    return role in (UserRole.FINANCE_TEAM, UserRole.FINANCE_CFA_TEAM,
                    UserRole.TENANT_ADMIN, UserRole.SUPER_ADMIN)


def style_ws(ws, headers: list, col_widths: list = None):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
        w = col_widths[col_idx - 1] if (col_widths and col_idx <= len(col_widths)) else 18
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 28


def style_data_cell(cell, alt: bool = False, status_color: str = None):
    cell.border = CELL_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if status_color:
        cell.fill = PatternFill("solid", fgColor=status_color)
        cell.font = Font(color="FFFFFF", bold=True)
    elif alt:
        cell.fill = ALT_FILL


def make_excel_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def get_role_filtered_claims(db, tenant_id, user: User,
                              status_filter=None, date_from=None,
                              date_to=None, search=None):
    """
    Returns claims filtered by role access rules.
    """
    q = db.query(PDCNClaim).filter(PDCNClaim.tenant_id == tenant_id)
    role = user.role

    # Scope by role
    if role == UserRole.DEALER:
        q = q.filter(PDCNClaim.dealer_id == user.id)
    elif role == UserRole.CN_TEAM:
        q = q.filter(PDCNClaim.status.in_(CN_VISIBLE_STATUSES))
    elif role == UserRole.CFA_TEAM:
        q = q.filter(PDCNClaim.status.in_(CFA_VISIBLE_STATUSES))
    # Finance / Admin: no additional filter — see all

    if status_filter:
        q = q.filter(PDCNClaim.status == status_filter)
    if date_from:
        try:
            q = q.filter(PDCNClaim.created_at >= datetime.fromisoformat(date_from))
        except Exception:
            pass
    if date_to:
        try:
            q = q.filter(PDCNClaim.created_at <= datetime.fromisoformat(date_to))
        except Exception:
            pass
    if search:
        q = q.filter(or_(
            PDCNClaim.claim_no.ilike("%" + search + "%"),
            PDCNClaim.dealer_name.ilike("%" + search + "%"),
        ))
    return q.order_by(PDCNClaim.created_at.desc()).all()


def get_remarks_for_role(claim: PDCNClaim, role: str) -> dict:
    """
    Returns only the remarks columns this role is allowed to see.
    Each role sees remarks from their own stage plus context.
    """
    remarks = {
        "cn_remarks":      "",
        "finance_remarks": "",
        "cfa_remarks":     "",
        "dealer_note":     "",
    }

    if is_finance(role):
        # Finance sees ALL remarks
        remarks["cn_remarks"]      = claim.cn_remarks or ""
        remarks["finance_remarks"] = claim.finance_remarks or ""
        remarks["cfa_remarks"]     = claim.cfa_remarks or ""

    elif role == UserRole.CN_TEAM:
        # CN sees only CN remarks (their own actions)
        remarks["cn_remarks"] = claim.cn_remarks or ""

    elif role == UserRole.CFA_TEAM:
        # CFA sees finance remarks (context) + CFA remarks (their own)
        remarks["finance_remarks"] = claim.finance_remarks or ""
        remarks["cfa_remarks"]     = claim.cfa_remarks or ""

    elif role == UserRole.DEALER:
        # Dealer sees all remarks for transparency on their claim
        remarks["cn_remarks"]      = claim.cn_remarks or ""
        remarks["finance_remarks"] = claim.finance_remarks or ""
        remarks["cfa_remarks"]     = claim.cfa_remarks or ""

    return remarks


# ─── Role-specific claims report ──────────────────────────────

@router.get("/claims/excel")
def export_claims_excel(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    role = current_user.role
    claims = get_role_filtered_claims(
        db, current_user.tenant_id, current_user,
        status, date_from, date_to, search
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    # Build headers based on role
    base_headers = [
        "Claim No", "Dealer Name", "Dealer Code", "Request Date",
        "Region", "State", "Sales Person", "Claim Type",
        "Total Qty", "Total Amount (₹)", "Status", "Submitted At",
    ]
    base_widths = [18, 22, 14, 14, 14, 14, 18, 16, 10, 18, 20, 16]

    # Remarks columns — role-specific
    if is_finance(role):
        ws.title = "PDCN Claims – Finance View"
        remark_headers = ["CN Remarks", "Finance Remarks", "CFA Remarks"]
        remark_widths  = [30, 30, 30]
    elif role == UserRole.CN_TEAM:
        ws.title = "PDCN Claims – CN View"
        remark_headers = ["CN Remarks"]
        remark_widths  = [36]
    elif role == UserRole.CFA_TEAM:
        ws.title = "PDCN Claims – CFA View"
        remark_headers = ["Finance Remarks", "CFA Remarks"]
        remark_widths  = [30, 30]
    elif role == UserRole.DEALER:
        ws.title = "My PDCN Claims"
        remark_headers = ["CN Remarks", "Finance Remarks", "CFA Remarks"]
        remark_widths  = [30, 30, 30]
    else:
        ws.title = "PDCN Claims"
        remark_headers = []
        remark_widths  = []

    headers = base_headers + remark_headers + ["Created At"]
    widths  = base_widths  + remark_widths  + [16]
    style_ws(ws, headers, widths)

    for i, c in enumerate(claims, 2):
        alt = i % 2 == 0
        remarks = get_remarks_for_role(c, role)

        base_row = [
            c.claim_no,
            c.dealer_name,
            c.dealer_code or "",
            c.request_date.strftime("%d-%m-%Y") if c.request_date else "",
            c.region or "",
            c.state or "",
            c.sales_person or "",
            (c.claim_type or "").upper(),
            c.total_qty,
            c.total_amount,
            (c.status or "").replace("_", " ").upper(),
            c.submitted_at.strftime("%d-%m-%Y") if c.submitted_at else "",
        ]

        if is_finance(role):
            remark_vals = [remarks["cn_remarks"], remarks["finance_remarks"], remarks["cfa_remarks"]]
        elif role == UserRole.CN_TEAM:
            remark_vals = [remarks["cn_remarks"]]
        elif role == UserRole.CFA_TEAM:
            remark_vals = [remarks["finance_remarks"], remarks["cfa_remarks"]]
        elif role == UserRole.DEALER:
            remark_vals = [remarks["cn_remarks"], remarks["finance_remarks"], remarks["cfa_remarks"]]
        else:
            remark_vals = []

        row = base_row + remark_vals + [
            c.created_at.strftime("%d-%m-%Y %H:%M") if c.created_at else ""
        ]

        status_col = 11  # Status is column 11
        for col_idx, val in enumerate(row, 1):
            status_color = None
            if col_idx == status_col:
                status_color = STATUS_COLORS.get((c.status or "").lower())
            cell = ws.cell(row=i, column=col_idx, value=val)
            style_data_cell(cell, alt, status_color)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    role_suffix = {
        UserRole.CN_TEAM: "CN", UserRole.CFA_TEAM: "CFA",
        UserRole.DEALER: "Dealer", UserRole.FINANCE_TEAM: "Finance",
        UserRole.FINANCE_CFA_TEAM: "Finance",
    }.get(role, "")
    fname = f"PDCN_Claims_{role_suffix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_excel_response(wb, fname)


# ─── Role-specific Line Items report ──────────────────────────

@router.get("/line-items/excel")
def export_line_items_excel(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    """Enhanced line-items export with full distributor info and all new columns."""
    from app.models import User as UserModel, Tenant
    role = current_user.role
    claims = get_role_filtered_claims(
        db, current_user.tenant_id, current_user, status, date_from, date_to
    )

    # Batch fetch dealer user info for distributor details
    dealer_ids = list({c.dealer_id for c in claims if c.dealer_id})
    dealers = {str(u.id): u for u in db.query(UserModel).filter(UserModel.id.in_(dealer_ids)).all()} if dealer_ids else {}

    # Get tenant (company) info
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    tenant_name = tenant.company_name if tenant else ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDCN Line Items"

    # ── Cover/Summary row ──
    BLUE_FILL  = PatternFill("solid", fgColor="1E1B4B")
    GOLD_FILL  = PatternFill("solid", fgColor="F59E0B")
    GREEN_FILL = PatternFill("solid", fgColor="059669")
    CYAN_FILL  = PatternFill("solid", fgColor="0891B2")

    # Distributor info columns + claim + line item columns
    dist_headers = [
        "S.No", "Distributor Name", "SAP Code", "Business Group",
        "Region", "City", "State", "Credit Check",
    ]
    claim_headers = [
        "Claim No", "Claim Status", "Request Date", "Claim Type",
        "Sales Person",
    ]
    item_headers = [
        "Material Code", "Brand Code", "Invoice No",
        "Manual Invoice?", "Invoice Date",
        "Product Name", "Product Code", "Batch No",
        "Quantity", "Purchase Price (₹)", "Allowed SD Price (₹)",
        "Selling Price (₹)", "Diff Price (₹)",
        "Billed To", "Billed Date", "Dealer Invoice No",
        "Credit Note Type", "Tax %", "Total Claim (₹)",
    ]
    remark_headers = []
    if is_finance(role):
        remark_headers = ["CN Remarks", "Finance Remarks", "CFA Remarks"]
    elif role == UserRole.CN_TEAM:
        remark_headers = ["CN Remarks"]
    elif role == UserRole.CFA_TEAM:
        remark_headers = ["Finance Remarks", "CFA Remarks"]
    elif role == UserRole.DEALER:
        remark_headers = ["CN Remarks", "Finance Remarks", "CFA Remarks"]

    all_headers = dist_headers + claim_headers + item_headers + remark_headers
    col_widths = [6,26,12,14,10,14,14,10,  16,18,12,16,14,  14,10,20,14,12,16,22,10,16,14,16,18,24,16,10,16,  24,24,24]
    # Trim or pad
    while len(col_widths) < len(all_headers): col_widths.append(16)

    # ── Header row with colour-coded groups ──
    for ci, h in enumerate(all_headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
        if ci <= len(dist_headers):
            cell.fill = PatternFill("solid", fgColor="1E1B4B")   # deep indigo – distributor
        elif ci <= len(dist_headers) + len(claim_headers):
            cell.fill = PatternFill("solid", fgColor="312E81")   # violet – claim
        elif ci <= len(dist_headers) + len(claim_headers) + len(item_headers):
            cell.fill = PatternFill("solid", fgColor="065F46")   # green – line item
        else:
            cell.fill = PatternFill("solid", fgColor="7C3AED")   # purple – remarks
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1] if ci<=len(col_widths) else 16

    ws.row_dimensions[1].height = 40

    # ── Company header block ──
    # We'll add a sub-header with role info below main headers
    row_idx = 2
    sno = 1
    for c in claims:
        remarks = get_remarks_for_role(c, role)
        dealer = dealers.get(str(c.dealer_id))
        items = c.line_items

        if not items:
            # Still include claim with blank item row
            items_to_use = [None]
        else:
            items_to_use = items

        for item in items_to_use:
            alt = row_idx % 2 == 0
            # Distributor columns
            dist_vals = [
                sno,
                (dealer.distributor_name or dealer.full_name) if dealer else c.dealer_name,
                dealer.sap_code or "" if dealer else c.dealer_code or "",
                dealer.business_group or "" if dealer else "",
                dealer.region or c.region or "" if dealer else c.region or "",
                dealer.city or "" if dealer else "",
                dealer.state or c.state or "" if dealer else c.state or "",
                dealer.credit_check or "" if dealer else "",
            ]
            # Claim columns
            claim_vals = [
                c.claim_no,
                (c.status or "").replace("_"," ").upper(),
                c.request_date.strftime("%d-%m-%Y") if c.request_date else "",
                (c.claim_type or "").upper(),
                c.sales_person or "",
            ]
            # Item columns
            if item:
                item_vals = [
                    item.sap_material_code or "",
                    item.brand_code or "",
                    item.invoice_no or "",
                    "Yes" if item.is_manual_invoice else "No",
                    item.invoice_date.strftime("%d-%m-%Y") if item.invoice_date else "",
                    item.product_name or "",
                    item.product_code or "",
                    item.batch_no or "",
                    item.quantity or 0,
                    item.purchase_price or "",
                    item.invoice_rate or "",
                    item.claim_rate or "",
                    item.difference_amt or "",
                    item.billed_to or "",
                    item.billed_date.strftime("%d-%m-%Y") if item.billed_date else "",
                    item.dealer_invoice_no or "",
                    item.credit_note_type or "CR1",
                    item.tax or "",
                    item.total_claim_amt or "",
                ]
            else:
                item_vals = [""] * len(item_headers)

            # Remark columns
            if is_finance(role):
                remark_vals = [remarks["cn_remarks"], remarks["finance_remarks"], remarks["cfa_remarks"]]
            elif role == UserRole.CN_TEAM:
                remark_vals = [remarks["cn_remarks"]]
            elif role == UserRole.CFA_TEAM:
                remark_vals = [remarks["finance_remarks"], remarks["cfa_remarks"]]
            elif role == UserRole.DEALER:
                remark_vals = [remarks["cn_remarks"], remarks["finance_remarks"], remarks["cfa_remarks"]]
            else:
                remark_vals = []

            all_vals = dist_vals + claim_vals + item_vals + remark_vals
            status_ci = len(dist_headers) + 2  # claim status column index

            for ci, val in enumerate(all_vals, 1):
                cell = ws.cell(row_idx, ci, val)
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(ci > len(dist_headers)+len(claim_headers)))
                if ci == status_ci:
                    sc = STATUS_COLORS.get((c.status or "").lower())
                    if sc:
                        cell.fill = PatternFill("solid", fgColor=sc)
                        cell.font = Font(color="FFFFFF", bold=True, size=9)
                        continue
                if alt:
                    cell.fill = ALT_FILL

            row_idx += 1
            sno += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["PDCN Line Items Export"])
    ws_sum.append(["Generated", datetime.utcnow().strftime("%d-%m-%Y %H:%M")])
    ws_sum.append(["Organisation", tenant_name])
    ws_sum.append(["Role", role.replace("_"," ").upper()])
    ws_sum.append(["Total Claims", len(claims)])
    ws_sum.append(["Total Line Items", row_idx - 2])
    ws_sum["A1"].font = Font(bold=True, size=14, color="1E1B4B")
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 30

    role_suffix = {
        UserRole.CN_TEAM:"CN", UserRole.CFA_TEAM:"CFA",
        UserRole.FINANCE_TEAM:"Finance", UserRole.FINANCE_CFA_TEAM:"Finance",
        UserRole.DEALER:"Dealer", UserRole.TENANT_ADMIN:"Admin",
    }.get(role, "")
    fname = f"PDCN_LineItems_{role_suffix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_excel_response(wb, fname)





# ─── Sales Register (Finance only) ────────────────────────────

@router.get("/sales-register/excel")
def export_sales_register_excel(
    search: Optional[str] = None,
    batch: Optional[str] = None,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if not is_finance(current_user.role):
        raise HTTPException(403, "Finance Team access required")

    q = db.query(SalesRegister).filter(
        SalesRegister.tenant_id == current_user.tenant_id
    )
    if search:
        q = q.filter(or_(
            SalesRegister.invoice_no.ilike("%" + search + "%"),
            SalesRegister.dealer_name.ilike("%" + search + "%"),
            SalesRegister.product_name.ilike("%" + search + "%"),
        ))
    if batch:
        q = q.filter(SalesRegister.upload_batch == batch)
    rows = q.order_by(SalesRegister.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Register"

    headers = ["Invoice No", "Invoice Date", "Dealer Name", "Dealer Code",
               "Product Name", "Product Code", "Batch No",
               "Quantity", "Rate (₹)", "Tax", "Net Amount (₹)", "Upload Batch"]
    widths  = [16, 14, 24, 14, 26, 14, 12, 10, 12, 8, 16, 14]
    style_ws(ws, headers, widths)

    for i, r in enumerate(rows, 2):
        alt = i % 2 == 0
        row = [
            r.invoice_no,
            r.invoice_date.strftime("%d-%m-%Y") if r.invoice_date else "",
            r.dealer_name, r.dealer_code or "",
            r.product_name, r.product_code or "", r.batch_no or "",
            r.quantity, r.rate, r.tax or "", r.net_amount or "", r.upload_batch,
        ]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            style_data_cell(cell, alt)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fname = f"SalesRegister_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_excel_response(wb, fname)


# ─── Qty Utilization (Finance only) ───────────────────────────

@router.get("/qty-utilization/excel")
def export_qty_utilization_excel(
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if not is_finance(current_user.role):
        raise HTTPException(403, "Finance Team access required")

    rows = db.query(QtyUtilizationLedger).filter(
        QtyUtilizationLedger.tenant_id == current_user.tenant_id
    ).order_by(QtyUtilizationLedger.invoice_no).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qty Utilization"

    headers = ["Invoice No", "Product Code", "Batch No",
               "Total Sold Qty", "Claimed Qty", "Balance Qty", "Updated At"]
    widths  = [18, 16, 14, 16, 14, 14, 20]
    style_ws(ws, headers, widths)

    for i, r in enumerate(rows, 2):
        alt = i % 2 == 0
        row = [r.invoice_no, r.product_code, r.batch_no or "",
               r.total_sold_qty, r.claimed_qty, r.balance_qty,
               r.updated_at.strftime("%d-%m-%Y %H:%M") if r.updated_at else ""]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            style_data_cell(cell, alt)
            if col_idx == 6 and isinstance(val, int) and val == 0:
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
                cell.font = Font(color="DC2626", bold=True)

    ws.freeze_panes = "A2"
    fname = f"QtyUtilization_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_excel_response(wb, fname)


# ─── Sales Register Template (Finance only) ───────────────────

@router.get("/template/excel")
def download_sales_register_template(
    current_user: User = Depends(get_active_tenant_user),
):
    if not is_finance(current_user.role):
        raise HTTPException(403, "Finance Team access required")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Register"

    headers = ["Invoice No", "Invoice Date", "Dealer Name", "Dealer Code",
               "Product Name", "Product Code", "Batch No",
               "Quantity", "Rate", "Tax", "Net Amount"]
    widths  = [18, 16, 26, 14, 28, 14, 14, 10, 12, 8, 16]
    style_ws(ws, headers, widths)

    for i in range(2, 22):
        alt = i % 2 == 0
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col_idx, value="")
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center")
            if alt:
                cell.fill = ALT_FILL

    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Column", "Required", "Format / Notes"),
        ("Invoice No",   "YES", "Unique invoice number — e.g. INV-2024-001"),
        ("Invoice Date", "YES", "YYYY-MM-DD — e.g. 2024-01-15"),
        ("Dealer Name",  "YES", "Full dealer name"),
        ("Dealer Code",  "No",  "Internal dealer code (optional)"),
        ("Product Name", "YES", "Full product name"),
        ("Product Code", "YES", "SKU or product code"),
        ("Batch No",     "No",  "Batch number (optional)"),
        ("Quantity",     "YES", "Integer units sold"),
        ("Rate",         "YES", "Per-unit rate e.g. 45.50"),
        ("Tax",          "No",  "Tax percentage e.g. 12"),
        ("Net Amount",   "No",  "Total invoice amount"),
    ]
    for row_data in instructions:
        ws2.append(row_data)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 50
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E1E3F")

    ws.freeze_panes = "A2"
    return make_excel_response(wb, "SalesRegister_Template.xlsx")


# ─── Single Claim Excel Export (exact format from reference) ───────────────

@router.get("/claims/{claim_id}/excel")
def export_single_claim_excel(
    claim_id: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    """Export single claim to Excel — exact reference format with 16-column layout."""
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from app.models import User as UserModel

    # ── Load & authorise ────────────────────────────────────────────────
    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id,
        PDCNClaim.tenant_id == current_user.tenant_id,
    ).first()
    if not claim:
        raise HTTPException(404, "Claim not found")
    if not can_view_claim(claim, current_user):
        raise HTTPException(403, "Access denied")

    dealer = db.query(UserModel).filter(UserModel.id == claim.dealer_id).first()
    items  = claim.line_items or []

    # ── Styles ──────────────────────────────────────────────────────────
    DARK_FILL  = PatternFill("solid", fgColor="1E1B4B")
    ALT_FILL   = PatternFill("solid", fgColor="F5F5FF")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    TOTAL_FILL = PatternFill("solid", fgColor="D9F0E8")
    APPROVED_FILL = PatternFill("solid", fgColor="C6EFCE")

    BOLD_WHITE = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    BOLD_DARK  = Font(bold=True, color="1E1B4B", name="Calibri", size=10)
    NORMAL     = Font(name="Calibri", size=10)
    ITALIC_SM  = Font(italic=True, name="Calibri", size=9)
    CLAIM_FONT = Font(bold=True, color="4338CA", name="Calibri", size=11)

    def thin():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Column widths (16 data cols + S.No) ────────────────────────────
    # Cols: S.No | Mat Code | Brand | Invoice No | Manual? | Inv Date |
    #        Product Name | Prod Code | Batch | Qty |
    #        Purchase Price(A) | Allowed SD(B) | Selling Price |
    #        Billed To | Billed Date | Dealer Inv No | CN Type
    # = 17 columns total
    COLS = 17
    widths = [5, 16, 10, 18, 10, 14, 28, 12, 10, 7, 16, 16, 16, 24, 14, 18, 12]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credit Note Claim"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Helper: dealer / claim info ────────────────────────────────────
    d_name  = (dealer.distributor_name or dealer.full_name or "") if dealer else ""
    d_sap   = (dealer.sap_code or "") if dealer else ""
    d_city  = (dealer.city or "")     if dealer else ""
    d_region= (dealer.region or "")   if dealer else ""
    d_state = (dealer.state or "")    if dealer else ""
    d_biz   = (dealer.business_group or "") if dealer else ""

    req_date = claim.request_date.strftime("%d-%m-%Y") if claim.request_date else ""
    claim_month = ""
    if claim.request_date:
        claim_month = claim.request_date.strftime("%B, %Y")

    # ── HEADER BLOCK (rows 1–6) ─────────────────────────────────────────
    def hrow(row, pairs, height=17):
        col = 1
        for label, value in pairs:
            ws.cell(row, col, label).font   = Font(bold=True, name="Calibri", size=10)
            ws.cell(row, col+1, value).font = BOLD_DARK if value else NORMAL
            col += 4
        ws.row_dimensions[row].height = height

    ws["A1"] = "Name of the Dealer";  ws["A1"].font = Font(bold=True, name="Calibri", size=10)
    ws["B1"] = d_name;                ws["B1"].font = BOLD_DARK
    ws["E1"] = "Customer Code (SAP)"; ws["E1"].font = Font(bold=True, name="Calibri", size=10)
    ws["F1"] = d_sap;                 ws["F1"].font = BOLD_DARK
    ws.row_dimensions[1].height = 17

    ws["A2"] = "Approval No";  ws["A2"].font = Font(bold=True, name="Calibri", size=10)
    ws["B2"] = claim.claim_no; ws["B2"].font = CLAIM_FONT
    ws["E2"] = "Business Group"; ws["E2"].font = Font(bold=True, name="Calibri", size=10)
    ws["F2"] = d_biz;            ws["F2"].font = NORMAL

    ws["A3"] = "Place";    ws["A3"].font = Font(bold=True, name="Calibri", size=10)
    ws["B3"] = d_city;     ws["B3"].font = NORMAL
    ws["E3"] = "Region";   ws["E3"].font = Font(bold=True, name="Calibri", size=10)
    ws["F3"] = d_region;   ws["F3"].font = NORMAL

    ws["A4"] = "Request Date"; ws["A4"].font = Font(bold=True, name="Calibri", size=10)
    ws["B4"] = req_date;       ws["B4"].font = NORMAL
    ws["E4"] = "State";        ws["E4"].font = Font(bold=True, name="Calibri", size=10)
    ws["F4"] = d_state;        ws["F4"].font = NORMAL

    ws["A5"] = "Claim for Month"; ws["A5"].font = Font(bold=True, name="Calibri", size=10)
    ws["B5"] = claim_month;       ws["B5"].font = NORMAL

    ws.row_dimensions[6].height = 8  # spacer

    # ── TABLE HEADERS (row 7) — exactly as requested ───────────────────
    HDR_ROW = 7
    headers = [
        "S. No",
        "Material Code",
        "Brand Code",
        "Invoice No",
        "If Manual Invoice? (yes/no)",
        "Invoice Date (DD-MM-YYYY)",
        "Product Name",
        "Product Code",
        "Batch No",
        "Quantity",
        "Purchase Price (A)",
        "Allowed SD Price (B)",
        "Selling Price (Dealer Invoice)",
        "Billed To",
        "Billed Date (DD-MM-YYYY)",
        "Dealer Invoice Number",
        "Credit Note Type (CR1/CR2)",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(HDR_ROW, ci, h)
        cell.font      = BOLD_WHITE
        cell.fill      = DARK_FILL
        cell.border    = thin()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[HDR_ROW].height = 42

    # ── DATA ROWS ────────────────────────────────────────────────────────
    # Diff formula: (A - B) * C  =  (Purchase Price - Allowed SD Price) * Qty
    # Total Claim = Sum of Diff across rows
    data_start = HDR_ROW + 1
    total_claim = 0.0

    for idx, item in enumerate(items):
        row = data_start + idx
        try:
            purch = float(item.purchase_price or 0)
            sd    = float(item.invoice_rate   or 0)   # Allowed SD Price
            qty   = int(item.quantity or 1)
        except Exception:
            purch, sd, qty = 0.0, 0.0, 1

        diff  = round((purch - sd) * qty, 2)   # (A-B)*C
        total_claim += diff

        inv_date  = item.invoice_date.strftime("%d-%m-%Y") if item.invoice_date  else ""
        bill_date = item.billed_date.strftime("%d-%m-%Y")  if item.billed_date   else ""
        manual    = "yes" if item.is_manual_invoice else "no"
        sell_price = float(item.claim_rate or 0)

        vals = [
            idx + 1,                              # S.No
            item.sap_material_code or "",         # Material Code
            item.brand_code or "",                # Brand Code
            item.invoice_no or "",                # Invoice No
            manual,                               # Manual Invoice?
            inv_date,                             # Invoice Date
            item.product_name or "",              # Product Name
            item.product_code or "",              # Product Code
            item.batch_no or "",                  # Batch No
            qty,                                  # Quantity
            purch if purch else "",               # Purchase Price (A)
            sd    if sd    else "",               # Allowed SD Price (B)
            sell_price if sell_price else "",     # Selling Price
            item.billed_to or "",                 # Billed To
            bill_date,                            # Billed Date
            item.dealer_invoice_no or "",         # Dealer Invoice Number
            item.credit_note_type or "CR1",       # CN Type
        ]
        fill = WHITE_FILL if idx % 2 == 0 else ALT_FILL
        RIGHT  = {11, 12, 13}   # Purchase, SD, Selling — right aligned
        CENTER = {1, 5, 6, 10, 15}  # S.No, Manual, Dates, Qty
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row, ci, v)
            cell.font   = NORMAL
            cell.fill   = fill
            cell.border = thin()
            if ci in RIGHT:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in CENTER:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Blank spacer rows
    last_data = data_start + len(items)
    for i in range(3):
        r = last_data + i
        for ci in range(1, COLS + 1):
            ws.cell(r, ci).border = thin()
            ws.cell(r, ci).fill   = WHITE_FILL

    # ── TOTALS SECTION ───────────────────────────────────────────────────
    tot_row = last_data + 4
    ws.row_dimensions[tot_row].height = 18

    total_claim_r = round(total_claim, 2)

    for tr, label, sym, val, is_approved in [
        (tot_row,     "Total Credit Note Value Claimed (INR)", "(X)",        total_claim_r, False),
        (tot_row + 2, "Less: Deduction / Adjustment (if any)", "(Y)",        "",             False),
        (tot_row + 4, "Total Credit Note Value Approved (INR)", "Z = (X-Y)", total_claim_r, True),
    ]:
        lc = ws.cell(tr, 13, label)
        lc.font      = BOLD_DARK
        lc.alignment = Alignment(horizontal="right")
        sc = ws.cell(tr, 15, sym)
        sc.font      = BOLD_DARK
        sc.alignment = Alignment(horizontal="center")
        vc = ws.cell(tr, 16, val if val != "" else "-")
        vc.font      = Font(bold=True, color="059669", name="Calibri", size=11) if val else NORMAL
        vc.fill      = APPROVED_FILL if is_approved else TOTAL_FILL
        vc.border    = thin()
        vc.alignment = Alignment(horizontal="right")

    # ── DECLARATION ───────────────────────────────────────────────────────
    decl_row = tot_row + 7
    ws.merge_cells(start_row=decl_row, start_column=1, end_row=decl_row, end_column=COLS)
    dc = ws.cell(decl_row, 1,
        "Distributor Confirmation — I hereby confirm that the data provided above in the "
        "credit note claim form and the supporting documents submitted are true and accurate.")
    dc.font      = ITALIC_SM
    dc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[decl_row].height = 30

    # ── SIGNATURES ────────────────────────────────────────────────────────
    sig_row = decl_row + 6
    ws.cell(sig_row,     1,  "Distributor Signature").font  = BOLD_DARK
    ws.cell(sig_row + 1, 1,  "(Authorised Signatory)").font = ITALIC_SM
    ws.cell(sig_row,     8,  "Approved By").font             = BOLD_DARK
    ws.cell(sig_row + 1, 8,  "(Sales / Marketing)").font    = ITALIC_SM
    ws.cell(sig_row,     13, "Approved By").font             = BOLD_DARK
    ws.cell(sig_row + 1, 13, "(Finance)").font               = ITALIC_SM

    # ── NOTE ──────────────────────────────────────────────────────────────
    note_row = sig_row + 4
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=COLS)
    nc = ws.cell(note_row, 1,
        "Note: Please attach copies of all sales invoices and purchase invoices "
        "for processing the Credit Note.")
    nc.font      = Font(italic=True, color="666666", size=8, name="Calibri")
    nc.alignment = Alignment(wrap_text=True)

    # ── Freeze & print ────────────────────────────────────────────────────
    ws.freeze_panes = f"A{HDR_ROW + 1}"
    ws.print_area   = f"A1:{get_column_letter(COLS)}{note_row}"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe = claim.claim_no.replace("/", "-").replace(" ", "_")
    fname = f"{safe}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

