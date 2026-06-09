"""
PDCN Claim Management API
Complete workflow: Dealer → CN Team → Finance → CFA → Completed
"""
import os
import uuid
import shutil
from datetime import datetime
from typing import Optional, List
from pathlib import Path

from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Form, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_

from app.auth.dependencies import get_active_tenant_user, get_tenant_admin
from app.database import get_db
from app.models import (
    User, UserRole, UserStatus,
    PDCNClaim, PDCNLineItem, PDCNAttachment, PDCNApprovalLog,
    CreditNote, SalesRegister, SalesRegisterBatch, QtyUtilizationLedger, PDCNNotification,
    PDCNStatus, Tenant
)

router = APIRouter(prefix="/pdcn", tags=["PDCN"])

UPLOAD_DIR = Path("uploads/pdcn")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

def _parse_date(val):
    """Parse date from multiple formats."""
    if not val:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('none', 'null', '—', '-'):
        return None
    from datetime import datetime as _dt
    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y',
                '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f']:
        try:
            return _dt.strptime(s, fmt)
        except ValueError:
            continue
    return None

ALLOWED_MIME = {"application/pdf", "image/jpeg", "image/png",
                "application/vnd.ms-excel",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


# ─── Helpers ──────────────────────────────────────────────────

def get_ip(request: Request):
    return request.client.host if request.client else "unknown"


def generate_claim_no(db: Session, tenant_id, dealer_user=None) -> str:
    """Generate claim number using dealer series if set.
    Format: {SERIES}-{Mon}-{YY}-{SEQ:03d}  e.g. MAX-SD-May-26-001
    Fallback: PDCN-{YYYYMM}-{SEQ:04d}
    """
    import random, string
    now = datetime.utcnow()
    month_abbr = now.strftime("%b").upper()   # MAY
    year_2     = now.strftime("%y")            # 26
    
    series = None
    if dealer_user and hasattr(dealer_user, 'dealer_series') and dealer_user.dealer_series:
        series = dealer_user.dealer_series.strip().upper()

    for _ in range(5):
        if series:
            # Count claims by this specific dealer
            count = db.query(PDCNClaim).filter(
                PDCNClaim.dealer_id == dealer_user.id,
                PDCNClaim.tenant_id == tenant_id,
            ).count()
            candidate = f"{series}-SD-{month_abbr}-{year_2}-{str(count + 1).zfill(3)}"
        else:
            count = db.query(PDCNClaim).filter(PDCNClaim.tenant_id == tenant_id).count()
            candidate = "PDCN-" + now.strftime("%Y%m") + "-" + str(count + 1).zfill(4)

        existing = db.query(PDCNClaim).filter(
            PDCNClaim.tenant_id == tenant_id,
            PDCNClaim.claim_no == candidate
        ).first()
        if not existing:
            return candidate
        # Collision – retry with bumped count

    # Final fallback
    return (series or "PDCN") + "-" + now.strftime("%Y%m%d%H%M%S") + "-" +            "".join(random.choices(string.digits, k=3))


def add_approval_log(db, claim, action, from_status, to_status, user, remarks, ip):
    log = PDCNApprovalLog(
        claim_id=claim.id,
        action=action,
        from_status=from_status,
        to_status=to_status,
        actioned_by=user.id,
        actioned_by_name=user.full_name,
        actioned_by_role=user.role,
        remarks=remarks,
        ip_address=ip,
    )
    db.add(log)


def notify(db, tenant_id, user_id, claim_id, title, message):
    n = PDCNNotification(
        tenant_id=tenant_id, user_id=user_id, claim_id=claim_id,
        title=title, message=message
    )
    db.add(n)


def claim_to_dict(c: PDCNClaim, include_items=False, include_logs=False):
    d = {
        "id": str(c.id),
        "claim_no": c.claim_no,
        "dealer_id": str(c.dealer_id),
        "dealer_name": c.dealer_name,
        "dealer_code": c.dealer_code,
        "request_date": c.request_date.isoformat() if c.request_date else None,
        "region": c.region,
        "state": c.state,
        "sales_person": c.sales_person,
        "claim_type": c.claim_type,
        "status": c.status,
        "total_qty": c.total_qty,
        "total_amount": c.total_amount,
        "cn_remarks": c.cn_remarks,
        "finance_remarks": c.finance_remarks,
        "cfa_remarks": c.cfa_remarks,
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "submitted_at": c.submitted_at.isoformat() if c.submitted_at else None,
        "attachments": [{"id": str(a.id), "original_name": a.original_name, "filename": a.filename, "file_size": a.file_size} for a in c.attachments],
        "credit_note": None,
    }
    if c.credit_note:
        cn = c.credit_note
        d["credit_note"] = {
            "id": str(cn.id), "cn_number": cn.cn_number,
            "cn_date": cn.cn_date.isoformat() if cn.cn_date else None,
            "cn_amount": cn.cn_amount, "cn_filename": cn.cn_filename,
            "cn_original_name": cn.cn_original_name,
        }
    if include_items:
        d["line_items"] = [{
            "id": str(i.id),
            "sap_material_code": i.sap_material_code or "",
            "brand_code": i.brand_code or "",
            "invoice_no": i.invoice_no,
            "is_manual_invoice": i.is_manual_invoice or False,
            "invoice_date": i.invoice_date.isoformat() if i.invoice_date else None,
            "product_name": i.product_name,
            "product_code": i.product_code or "",
            "batch_no": i.batch_no or "",
            "quantity": i.quantity,
            "purchase_price": i.purchase_price or "",
            "invoice_rate": i.invoice_rate,
            "claim_rate": i.claim_rate,
            "difference_amt": i.difference_amt or "",
            "billed_to": i.billed_to or "",
            "billed_date": i.billed_date.isoformat() if i.billed_date else None,
            "dealer_invoice_no": i.dealer_invoice_no or "",
            "credit_note_type": i.credit_note_type or "CR1",
            "tax": i.tax or "",
            "total_claim_amt": i.total_claim_amt or "",
            "reason": i.reason or "",
            "remarks": i.remarks or "",
        } for i in c.line_items]
    if include_logs:
        d["approval_logs"] = [{
            "id": str(l.id), "action": l.action,
            "from_status": l.from_status, "to_status": l.to_status,
            "actioned_by_name": l.actioned_by_name, "actioned_by_role": l.actioned_by_role,
            "remarks": l.remarks, "created_at": l.created_at.isoformat() if l.created_at else None,
        } for l in sorted(c.approval_logs, key=lambda x: x.created_at)]
    return d


def can_view_claim(claim: PDCNClaim, user: User) -> bool:
    """Check if user can view a specific claim. Drafts are private to dealer/admin."""
    role = user.role
    if role == UserRole.SUPER_ADMIN:
        return True
    if role == UserRole.TENANT_ADMIN:
        return True  # Admin sees everything in their tenant
    if role == UserRole.DEALER:
        return str(claim.dealer_id) == str(user.id)
    # All other roles cannot see DRAFT claims (only dealer/admin can)
    if claim.status == PDCNStatus.DRAFT:
        return False
    if role == UserRole.CN_TEAM:
        return True  # CN sees submitted and above
    if role == UserRole.FINANCE_TEAM:
        return True  # Finance sees submitted and above
    if role == UserRole.CFA_TEAM:
        return claim.status in (
            PDCNStatus.FINANCE_APPROVED, PDCNStatus.CN_PENDING,
            PDCNStatus.CN_GENERATED, PDCNStatus.COMPLETED
        )
    if role == UserRole.FINANCE_CFA_TEAM:
        return True
    return False


def get_visible_claims(q, user: User):
    """Filter claims based on role — draft claims only visible to dealer & admin."""
    role = user.role
    if role == UserRole.DEALER:
        # Dealers only see their own claims
        return q.filter(PDCNClaim.dealer_id == user.id)
    if role == UserRole.CFA_TEAM:
        # CFA sees only from finance_approved onwards
        return q.filter(PDCNClaim.status.in_([
            PDCNStatus.FINANCE_APPROVED, PDCNStatus.CN_PENDING,
            PDCNStatus.CN_GENERATED, PDCNStatus.COMPLETED
        ]))
    if role in (UserRole.CN_TEAM, UserRole.FINANCE_TEAM):
        # CN and Finance teams see all EXCEPT draft
        return q.filter(PDCNClaim.status != PDCNStatus.DRAFT)
    if role == UserRole.FINANCE_CFA_TEAM:
        # Finance+CFA sees all except draft
        return q.filter(PDCNClaim.status != PDCNStatus.DRAFT)
    # Tenant Admin and Super Admin see everything
    return q


# ─── Dashboard Stats ───────────────────────────────────────────

@router.get("/dashboard")
def pdcn_dashboard(
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    tid = current_user.tenant_id
    base = db.query(PDCNClaim).filter(PDCNClaim.tenant_id == tid)
    base = get_visible_claims(base, current_user)

    def count(status_val):
        return base.filter(PDCNClaim.status == status_val).count()

    total = base.count()
    return {
        "total":             total,
        "draft":             count(PDCNStatus.DRAFT),
        "submitted":         count(PDCNStatus.SUBMITTED),
        "cn_approved":       count(PDCNStatus.CN_APPROVED),
        "cn_rejected":       count(PDCNStatus.CN_REJECTED),
        "finance_approved":  count(PDCNStatus.FINANCE_APPROVED),
        "finance_rejected":  count(PDCNStatus.FINANCE_REJECTED),
        "sent_back":         count(PDCNStatus.SENT_BACK),
        "cn_pending":        count(PDCNStatus.CN_PENDING),
        "cn_generated":      count(PDCNStatus.CN_GENERATED),
        "completed":         count(PDCNStatus.COMPLETED),
        "cancelled":         count(PDCNStatus.CANCELLED),
    }


# ─── List Claims ───────────────────────────────────────────────

@router.get("/claims")
def list_claims(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    dealer_id: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    q = db.query(PDCNClaim).filter(PDCNClaim.tenant_id == current_user.tenant_id)
    q = get_visible_claims(q, current_user)

    if status:
        q = q.filter(PDCNClaim.status == status)
    if dealer_id:
        q = q.filter(PDCNClaim.dealer_id == dealer_id)
    if search:
        q = q.filter(or_(
            PDCNClaim.claim_no.ilike("%" + search + "%"),
            PDCNClaim.dealer_name.ilike("%" + search + "%"),
        ))
    if date_from:
        try:
            q = q.filter(PDCNClaim.created_at >= datetime.fromisoformat(date_from))
        except Exception:
            pass
    if date_to:
        try:
            q = q.filter(PDCNClaim.created_at <= datetime.fromisoformat(date_to))
        except Exception:
            pass

    total = q.count()
    claims = q.order_by(PDCNClaim.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "claims": [claim_to_dict(c) for c in claims],
        "total": total, "page": page, "per_page": per_page,
        "pages": max(1, (total + per_page - 1) // per_page)
    }


# ─── Get Claim Detail ──────────────────────────────────────────

@router.get("/claims/{claim_id}")
def get_claim(
    claim_id: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id,
        PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if not can_view_claim(claim, current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    return claim_to_dict(claim, include_items=True, include_logs=True)



@router.delete("/claims/{claim_id}")
def delete_claim(
    claim_id: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    """Permanently delete a DRAFT claim. Only the dealer who owns it can delete."""
    import uuid as _uuid
    try:
        claim_uuid = _uuid.UUID(str(claim_id))
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid claim ID format")

    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_uuid,
        PDCNClaim.tenant_id == current_user.tenant_id,
    ).first()
    if not claim:
        raise HTTPException(404, "Claim not found")

    # Only draft or sent_back can be deleted
    if claim.status not in (PDCNStatus.DRAFT, PDCNStatus.SENT_BACK):
        raise HTTPException(400, f"Only Draft claims can be deleted. This claim is '{claim.status}'.")

    # Only the owning dealer (or admin) can delete
    if current_user.role == UserRole.DEALER and str(claim.dealer_id) != str(current_user.id):
        raise HTTPException(403, "You can only delete your own claims.")

    # Cascade delete line items, attachments, approval logs
    from app.models import PDCNApprovalLog, PDCNAttachment, PDCNNotification
    db.query(PDCNLineItem).filter(PDCNLineItem.claim_id == claim.id).delete()
    db.query(PDCNApprovalLog).filter(PDCNApprovalLog.claim_id == claim.id).delete()
    db.query(PDCNAttachment).filter(PDCNAttachment.claim_id == claim.id).delete()
    db.query(PDCNNotification).filter(PDCNNotification.claim_id == claim.id).delete()
    db.delete(claim)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, "Failed to delete claim: " + str(e))
    return {"message": "Claim deleted successfully."}


# ─── Create Claim (Dealer) ─────────────────────────────────────

class ClaimCreateBody(BaseModel):
    dealer_name: str
    dealer_code: str = ""
    region: str = ""
    state: str = ""
    sales_person: str = ""
    claim_type: str = "pdcn"
    claim_month: str = ""   # e.g. "May"
    claim_year: str = ""    # e.g. "2026"
    items: list = []


@router.post("/claims", status_code=201)
def create_claim(
    request: Request,
    body: ClaimCreateBody,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    dealer_name = body.dealer_name
    dealer_code = body.dealer_code
    region = body.region
    state = body.state
    sales_person = body.sales_person
    claim_type = body.claim_type
    if current_user.role not in (UserRole.DEALER, UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="Only dealers can create claims")

    # ── Gate: Sales Register for the specific claim month must be uploaded ──
    if current_user.role == UserRole.DEALER:
        # Parse claim month/year from request body
        now = datetime.utcnow()
        claim_month_num = now.month
        claim_year_num  = now.year
        month_map = {
            "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
            "july":7,"august":8,"september":9,"october":10,"november":11,"december":12
        }
        if body.claim_month:
            claim_month_num = month_map.get(body.claim_month.lower().strip(), now.month)
        if body.claim_year:
            try: claim_year_num = int(body.claim_year)
            except Exception: pass

        # Build the batch_key for the SPECIFIC month being claimed
        # (can be current month OR any past month — dealer decides)
        batch_key = f"{claim_year_num:04d}-{claim_month_num:02d}"

        # Check if Finance has uploaded the SR for that exact month
        sr_batch = db.query(SalesRegisterBatch).filter(
            SalesRegisterBatch.tenant_id == current_user.tenant_id,
            SalesRegisterBatch.batch_key == batch_key,
        ).first()

        if not sr_batch:
            # ── Bypass: if ALL submitted items are manual invoices, allow claim ──
            all_manual = bool(body.items) and all(
                bool(item.get("is_manual_invoice") or item.get("manual_invoice"))
                for item in body.items
            )
            if all_manual:
                # All line items are manual — skip SR gate entirely
                pass
            else:
                month_name = datetime(claim_year_num, claim_month_num, 1).strftime("%B %Y")
                # Check if any items are non-manual (mixed scenario)
                has_non_manual = any(
                    not (item.get("is_manual_invoice") or item.get("manual_invoice"))
                    for item in body.items
                )
                # Fetch all uploaded months to guide the dealer
                uploaded = db.query(SalesRegisterBatch).filter(
                    SalesRegisterBatch.tenant_id == current_user.tenant_id,
                ).order_by(SalesRegisterBatch.batch_key.desc()).all()
                if uploaded:
                    available = ", ".join(
                        b.month_label or datetime.strptime(b.batch_key, "%Y-%m").strftime("%B %Y")
                        for b in uploaded
                    )
                    if has_non_manual:
                        detail = (
                            f"Sales register for {month_name} has not been uploaded yet. "
                            f"Some line items are not marked as manual. "
                            f"Either mark ALL line items as manual invoices to bypass this check, "
                            f"or select an available month ({available}), "
                            f"or ask Finance to upload the sales register for {month_name}."
                        )
                    else:
                        detail = (
                            f"Sales register for {month_name} has not been uploaded yet. "
                            f"Available months: {available}. "
                            f"Please select one of these months, ask Finance to upload {month_name}, "
                            f"or mark all line items as Manual Invoices to proceed without a sales register."
                        )
                else:
                    detail = (
                        f"No sales register has been uploaded by the Finance Team yet. "
                        f"You can still submit this claim by marking ALL line items as Manual Invoices."
                    )
                raise HTTPException(status_code=400, detail=detail)

    claim_no = generate_claim_no(db, current_user.tenant_id, current_user)
    claim = PDCNClaim(
        tenant_id=current_user.tenant_id,
        claim_no=claim_no,
        dealer_id=current_user.id,
        dealer_name=dealer_name,
        dealer_code=dealer_code,
        region=region,
        state=state,
        sales_person=sales_person,
        claim_type=claim_type,
        status=PDCNStatus.DRAFT,
        created_by=current_user.id,
    )
    try:
        db.add(claim)
        db.commit()
        db.refresh(claim)
        add_approval_log(db, claim, "CREATED", None, PDCNStatus.DRAFT, current_user, "Claim created", get_ip(request))
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed to create claim: " + str(e))

    # Also save any items submitted with the claim
    if body.items:
        for item in body.items:
            li = PDCNLineItem(
                claim_id=claim.id,
                sap_material_code=item.get("sap_material_code") or None,
                brand_code=item.get("brand_code") or None,
                invoice_no=item.get("invoice_no") or "MANUAL",
                is_manual_invoice=bool(item.get("is_manual_invoice", False)),
                invoice_date=_parse_date(item.get("invoice_date")),
                product_name=item.get("product_name") or "",
                product_code=item.get("product_code") or None,
                batch_no=item.get("batch_no") or None,
                quantity=int(item.get("quantity") or 1),
                purchase_price=str(item.get("purchase_price") or "0"),
                invoice_rate=str(item.get("invoice_rate") or "0"),
                claim_rate=str(item.get("claim_rate") or "0"),
                difference_amt=str(item.get("difference_amt") or "0"),
                billed_to=item.get("billed_to") or None,
                billed_date=_parse_date(item.get("billed_date")),
                dealer_invoice_no=item.get("dealer_invoice_no") or None,
                credit_note_type=item.get("credit_note_type") or "CR1",
                total_claim_amt=str(item.get("total_claim_amt") or "0"),
            )
            db.add(li)
        total_qty = sum(int(i.get("quantity") or 1) for i in body.items)
        try:
            total_amt = sum(float(i.get("total_claim_amt") or 0) for i in body.items)
        except Exception:
            total_amt = 0
        claim.total_qty = total_qty
        claim.total_amount = str(round(total_amt, 2))
        try:
            db.commit()
        except Exception as e:
            db.rollback()

    return claim_to_dict(claim)


# ─── Add / Update Line Items ───────────────────────────────────

@router.post("/claims/{claim_id}/items")
def add_line_items(
    claim_id: str,
    items: List[dict],
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id,
        PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if str(claim.dealer_id) != str(current_user.id) and current_user.role not in (UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="Access denied")
    if claim.status not in (PDCNStatus.DRAFT, PDCNStatus.SENT_BACK):
        raise HTTPException(status_code=400, detail="Cannot edit claim in status: " + claim.status)

    # Remove old items and replace
    db.query(PDCNLineItem).filter(PDCNLineItem.claim_id == claim.id).delete()

    total_qty = 0
    total_amt = 0.0

    for item_data in items:
        inv_no      = str(item_data.get("invoice_no", "")).strip()
        prod_code   = str(item_data.get("product_code", "")).strip()
        batch_no    = str(item_data.get("batch_no", "")).strip()
        qty         = int(item_data.get("quantity", 0))

        # Validate against sales register
        sr = db.query(SalesRegister).filter(
            SalesRegister.tenant_id == current_user.tenant_id,
            SalesRegister.invoice_no == inv_no,
        ).first()

        if sr:
            # Qty utilization check
            ledger = db.query(QtyUtilizationLedger).filter(
                QtyUtilizationLedger.tenant_id == current_user.tenant_id,
                QtyUtilizationLedger.invoice_no == inv_no,
                QtyUtilizationLedger.product_code == (prod_code or sr.product_code),
            ).first()
            if ledger and qty > ledger.balance_qty:
                raise HTTPException(
                    status_code=400,
                    detail="Qty " + str(qty) + " exceeds balance qty " + str(ledger.balance_qty) + " for invoice " + inv_no
                )

        line = PDCNLineItem(
            claim_id=claim.id,
            sap_material_code=str(item_data.get("sap_material_code","")).strip() or None,
            brand_code=str(item_data.get("brand_code","")).strip() or None,
            invoice_no=inv_no,
            is_manual_invoice=bool(item_data.get("is_manual_invoice", False)),
            invoice_date=_parse_date(item_data.get("invoice_date")),
            product_name=str(item_data.get("product_name", "")),
            product_code=prod_code,
            batch_no=batch_no,
            quantity=qty,
            purchase_price=str(item_data.get("purchase_price","0")).strip() or "0",
            invoice_rate=str(item_data.get("invoice_rate", "0")),
            claim_rate=str(item_data.get("claim_rate", "0")),
            difference_amt=str(item_data.get("difference_amt", "0")),
            billed_to=str(item_data.get("billed_to","")).strip() or None,
            billed_date=_parse_date(item_data.get("billed_date")),
            dealer_invoice_no=str(item_data.get("dealer_invoice_no","")).strip() or None,
            credit_note_type=str(item_data.get("credit_note_type","CR1")).strip() or "CR1",
            tax=str(item_data.get("tax", "0")),
            total_claim_amt=str(item_data.get("total_claim_amt", "0")),
            reason=str(item_data.get("reason", "")),
            remarks=str(item_data.get("remarks", "")),
        )
        db.add(line)
        total_qty += qty
        try:
            total_amt += float(item_data.get("total_claim_amt", 0))
        except Exception:
            pass

    # Recalculate total from diff: sum of (purchase_price - invoice_rate) * qty
    recalc_total = 0.0
    for it in claim.line_items:
        try:
            p = float(it.purchase_price or 0)
            s = float(it.invoice_rate or 0)
            q = int(it.quantity or 1)
            recalc_total += (p - s) * q
        except Exception:
            pass
    claim.total_qty    = total_qty
    claim.total_amount = str(round(recalc_total if recalc_total else total_amt, 2))
    db.commit()
    return {"message": "Line items saved", "total_qty": total_qty, "total_amount": claim.total_amount}


# ─── Submit Claim ──────────────────────────────────────────────

@router.post("/claims/{claim_id}/submit")
def submit_claim(
    claim_id: str,
    request: Request,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id, PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if str(claim.dealer_id) != str(current_user.id) and current_user.role not in (UserRole.SUPER_ADMIN,):
        raise HTTPException(status_code=403, detail="Only claim owner can submit")
    if claim.status not in (PDCNStatus.DRAFT, PDCNStatus.SENT_BACK):
        raise HTTPException(status_code=400, detail="Claim cannot be submitted from status: " + claim.status)
    if not claim.line_items:
        raise HTTPException(status_code=400, detail="Please add at least one line item before submitting")

    old_status = claim.status
    claim.status = PDCNStatus.SUBMITTED
    claim.submitted_at = datetime.utcnow()
    add_approval_log(db, claim, "SUBMITTED", old_status, PDCNStatus.SUBMITTED, current_user, None, get_ip(request))

    # Update qty utilization ledger
    for item in claim.line_items:
        if not item.product_code:
            continue
        ledger = db.query(QtyUtilizationLedger).filter(
            QtyUtilizationLedger.tenant_id == current_user.tenant_id,
            QtyUtilizationLedger.invoice_no == item.invoice_no,
            QtyUtilizationLedger.product_code == item.product_code,
        ).first()
        if ledger:
            ledger.claimed_qty += item.quantity
            ledger.balance_qty -= item.quantity
            ledger.balance_qty = max(0, ledger.balance_qty)

    db.commit()
    return {"message": "Claim submitted successfully", "claim_no": claim.claim_no}


# ─── CN Team Actions ───────────────────────────────────────────

@router.post("/claims/{claim_id}/cn-action")
def cn_team_action(
    claim_id: str,
    request: Request,
    action: str = Form(...),   # approve | reject | send_back
    remarks: str = Form(""),
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in (UserRole.CN_TEAM, UserRole.FINANCE_CFA_TEAM, UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="CN Team access required")

    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id, PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if claim.status != PDCNStatus.SUBMITTED:
        raise HTTPException(status_code=400, detail="Claim must be in SUBMITTED status for CN action")

    old_status = claim.status
    if action == "approve":
        claim.status = PDCNStatus.CN_APPROVED
        claim.cn_remarks = remarks
        claim.cn_actioned_by = current_user.id
        claim.cn_actioned_at = datetime.utcnow()
        add_approval_log(db, claim, "CN_APPROVED", old_status, PDCNStatus.CN_APPROVED, current_user, remarks, get_ip(request))
    elif action == "reject":
        if not remarks:
            raise HTTPException(status_code=400, detail="Remarks required for rejection")
        claim.status = PDCNStatus.CN_REJECTED
        claim.cn_remarks = remarks
        claim.cn_actioned_by = current_user.id
        claim.cn_actioned_at = datetime.utcnow()
        add_approval_log(db, claim, "CN_REJECTED", old_status, PDCNStatus.CN_REJECTED, current_user, remarks, get_ip(request))
    elif action == "send_back":
        if not remarks:
            raise HTTPException(status_code=400, detail="Remarks required for send back")
        claim.status = PDCNStatus.SENT_BACK
        claim.cn_remarks = remarks
        claim.cn_actioned_by = current_user.id
        claim.cn_actioned_at = datetime.utcnow()
        add_approval_log(db, claim, "SENT_BACK", old_status, PDCNStatus.SENT_BACK, current_user, remarks, get_ip(request))
    else:
        raise HTTPException(status_code=400, detail="Invalid action: " + action)

    db.commit()
    return {"message": "Action completed", "new_status": claim.status}


# ─── Finance Team Actions ──────────────────────────────────────

@router.post("/claims/{claim_id}/finance-action")
def finance_action(
    claim_id: str,
    request: Request,
    action: str = Form(...),   # approve | reject
    remarks: str = Form(""),
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in (UserRole.FINANCE_TEAM, UserRole.FINANCE_CFA_TEAM, UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="Finance Team access required")

    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id, PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if claim.status != PDCNStatus.CN_APPROVED:
        raise HTTPException(status_code=400, detail="Claim must be CN_APPROVED for Finance action")

    old_status = claim.status
    if action == "approve":
        claim.status = PDCNStatus.FINANCE_APPROVED
        claim.finance_remarks = remarks
        claim.fin_actioned_by = current_user.id
        claim.fin_actioned_at = datetime.utcnow()
        add_approval_log(db, claim, "FINANCE_APPROVED", old_status, PDCNStatus.FINANCE_APPROVED, current_user, remarks, get_ip(request))
    elif action == "reject":
        if not remarks:
            raise HTTPException(status_code=400, detail="Remarks required for rejection")
        claim.status = PDCNStatus.FINANCE_REJECTED
        claim.finance_remarks = remarks
        claim.fin_actioned_by = current_user.id
        claim.fin_actioned_at = datetime.utcnow()
        add_approval_log(db, claim, "FINANCE_REJECTED", old_status, PDCNStatus.FINANCE_REJECTED, current_user, remarks, get_ip(request))
    else:
        raise HTTPException(status_code=400, detail="Invalid action: " + action)

    db.commit()
    return {"message": "Finance action completed", "new_status": claim.status}


# ─── CFA: Generate Credit Note ─────────────────────────────────

@router.post("/claims/{claim_id}/credit-note")
async def generate_credit_note(
    claim_id: str,
    request: Request,
    cn_number: str = Form(...),
    cn_date: str = Form(...),
    cn_amount: str = Form(...),
    remarks: str = Form(""),
    file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in (UserRole.CFA_TEAM, UserRole.FINANCE_CFA_TEAM, UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="CFA Team access required")

    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id, PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if claim.status != PDCNStatus.FINANCE_APPROVED:
        raise HTTPException(status_code=400, detail="Claim must be FINANCE_APPROVED to generate Credit Note")

    cn_filename = None
    cn_original_name = None
    if file:
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="File too large (max 10MB)")
        ext = Path(file.filename).suffix.lower()
        safe_name = str(uuid.uuid4()) + ext
        save_path = UPLOAD_DIR / safe_name
        with open(save_path, "wb") as f:
            f.write(content)
        cn_filename = safe_name
        cn_original_name = file.filename

    existing_cn = db.query(CreditNote).filter(CreditNote.claim_id == claim.id).first()
    if existing_cn:
        existing_cn.cn_number = cn_number
        existing_cn.cn_date = datetime.fromisoformat(cn_date)
        existing_cn.cn_amount = cn_amount
        existing_cn.remarks = remarks
        if cn_filename:
            existing_cn.cn_filename = cn_filename
            existing_cn.cn_original_name = cn_original_name
    else:
        cn = CreditNote(
            claim_id=claim.id,
            tenant_id=current_user.tenant_id,
            cn_number=cn_number,
            cn_date=datetime.fromisoformat(cn_date),
            cn_amount=cn_amount,
            cn_filename=cn_filename,
            cn_original_name=cn_original_name,
            remarks=remarks,
            created_by=current_user.id,
        )
        db.add(cn)

    old_status = claim.status
    claim.status = PDCNStatus.CN_GENERATED
    claim.cfa_actioned_by = current_user.id
    claim.cfa_actioned_at = datetime.utcnow()
    claim.cfa_remarks = remarks
    add_approval_log(db, claim, "CN_GENERATED", old_status, PDCNStatus.CN_GENERATED, current_user, remarks, get_ip(request))

    # Notify dealer
    notify(db, claim.tenant_id, claim.dealer_id, claim.id,
           "Credit Note Generated",
           "Credit Note " + cn_number + " has been generated for your claim " + claim.claim_no)

    db.commit()
    return {"message": "Credit Note generated", "cn_number": cn_number}


# ─── CFA: Mark Completed ───────────────────────────────────────

@router.post("/claims/{claim_id}/complete")
def complete_claim(
    claim_id: str,
    request: Request,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in (UserRole.CFA_TEAM, UserRole.FINANCE_CFA_TEAM, UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="CFA Team access required")

    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id, PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if claim.status != PDCNStatus.CN_GENERATED:
        raise HTTPException(status_code=400, detail="Claim must have CN_GENERATED to complete")

    old_status = claim.status
    claim.status = PDCNStatus.COMPLETED
    add_approval_log(db, claim, "COMPLETED", old_status, PDCNStatus.COMPLETED, current_user, None, get_ip(request))
    db.commit()
    return {"message": "Claim completed successfully"}


# ─── File Upload (Attachments) ─────────────────────────────────

@router.post("/claims/{claim_id}/attachments")
async def upload_attachment(
    claim_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id, PDCNClaim.tenant_id == current_user.tenant_id
    ).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    if claim.status not in (PDCNStatus.DRAFT, PDCNStatus.SENT_BACK):
        raise HTTPException(status_code=400, detail="Cannot upload to claim in status: " + claim.status)

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")

    ext = Path(file.filename).suffix.lower()
    safe_name = str(uuid.uuid4()) + ext
    save_path = UPLOAD_DIR / safe_name
    with open(save_path, "wb") as f:
        f.write(content)

    att = PDCNAttachment(
        claim_id=claim.id,
        filename=safe_name,
        original_name=file.filename,
        file_size=len(content),
        mime_type=file.content_type,
        uploaded_by=current_user.id,
    )
    db.add(att)
    db.commit()
    return {"id": str(att.id), "original_name": att.original_name, "filename": att.filename}


@router.delete("/claims/{claim_id}/attachments/{att_id}")
def delete_attachment(
    claim_id: str,
    att_id: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    att = db.query(PDCNAttachment).filter(
        PDCNAttachment.id == att_id, PDCNAttachment.claim_id == claim_id
    ).first()
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    try:
        (UPLOAD_DIR / att.filename).unlink(missing_ok=True)
    except Exception:
        pass
    db.delete(att)
    db.commit()
    return {"message": "Attachment deleted"}


@router.get("/attachments/{filename}")
def download_attachment(
    filename: str,
    current_user: User = Depends(get_active_tenant_user),
):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(str(file_path))


# ─── Sales Register Upload ─────────────────────────────────────

@router.post("/sales-register/upload")
async def upload_sales_register(
    file: UploadFile = File(...),
    upload_batch: str = Form(...),
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in (UserRole.FINANCE_TEAM, UserRole.FINANCE_CFA_TEAM, UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="Finance Team access required")

    content = await file.read()
    import io
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        try:
            import csv
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail="Invalid file format: " + str(e))

    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    # Auto-detect header row
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    def get(row, key, fallback=""):
        idx = col.get(key)
        if idx is None:
            return fallback
        v = row[idx] if idx < len(row) else None
        return str(v).strip() if v is not None else fallback

    inserted = 0
    skipped = 0
    for row in rows[1:]:
        if not any(row):
            continue
        inv_no = get(row, "invoice_no") or get(row, "invoice")
        if not inv_no:
            continue

        # Parse qty
        try:
            qty = int(float(get(row, "quantity") or get(row, "qty") or 0))
        except Exception:
            qty = 0

        # Parse date
        inv_date_raw = get(row, "invoice_date") or get(row, "date")
        try:
            if isinstance(inv_date_raw, datetime):
                inv_date = inv_date_raw
            else:
                inv_date = datetime.fromisoformat(str(inv_date_raw))
        except Exception:
            inv_date = datetime.utcnow()

        prod_code = get(row, "product_code")
        batch_no  = get(row, "batch_no") or get(row, "batch")

        # Check duplicate
        existing = db.query(SalesRegister).filter(
            SalesRegister.tenant_id == current_user.tenant_id,
            SalesRegister.invoice_no == inv_no,
            SalesRegister.product_code == prod_code,
            SalesRegister.batch_no == batch_no,
        ).first()
        if existing:
            skipped += 1
            continue

        sr = SalesRegister(
            tenant_id=current_user.tenant_id,
            upload_batch=upload_batch,
            invoice_no=inv_no,
            invoice_date=inv_date,
            dealer_name=get(row, "dealer_name") or get(row, "dealer"),
            dealer_code=get(row, "dealer_code"),
            product_name=get(row, "product_name") or get(row, "product"),
            product_code=prod_code,
            batch_no=batch_no,
            quantity=qty,
            rate=get(row, "rate") or get(row, "invoice_rate"),
            tax=get(row, "tax"),
            net_amount=get(row, "net_amount"),
            uploaded_by=current_user.id,
        )
        db.add(sr)

        # Upsert qty ledger
        ledger = db.query(QtyUtilizationLedger).filter(
            QtyUtilizationLedger.tenant_id == current_user.tenant_id,
            QtyUtilizationLedger.invoice_no == inv_no,
            QtyUtilizationLedger.product_code == prod_code,
        ).first()
        if not ledger:
            ledger = QtyUtilizationLedger(
                tenant_id=current_user.tenant_id,
                invoice_no=inv_no,
                product_code=prod_code,
                batch_no=batch_no,
                total_sold_qty=qty,
                claimed_qty=0,
                balance_qty=qty,
            )
            db.add(ledger)
        inserted += 1

    db.commit()

    # ── Record batch so dealers can create claims for this month ──────
    try:
        from datetime import datetime as _dt
        batch_key = upload_batch.strip()
        for fmt in ["%Y-%m", "%m-%Y", "%B %Y", "%b %Y", "%B-%Y"]:
            try:
                parsed = _dt.strptime(batch_key, fmt)
                batch_key = parsed.strftime("%Y-%m")
                break
            except ValueError:
                continue
        existing_b = db.query(SalesRegisterBatch).filter(
            SalesRegisterBatch.tenant_id == current_user.tenant_id,
            SalesRegisterBatch.batch_key == batch_key,
        ).first()
        if existing_b:
            existing_b.row_count   = inserted + skipped
            existing_b.uploaded_by = current_user.id
            existing_b.uploaded_at = _dt.utcnow()
        else:
            db.add(SalesRegisterBatch(
                tenant_id   = current_user.tenant_id,
                batch_key   = batch_key,
                month_label = upload_batch.strip(),
                uploaded_by = current_user.id,
                row_count   = inserted + skipped,
            ))
        db.commit()
        print(f"[INFO] SR batch recorded: {batch_key}")
    except Exception as be:
        db.rollback()
        print(f"[WARN] Could not record SR batch: {be}")

    return {"message": "Upload complete", "inserted": inserted, "skipped_duplicates": skipped}


@router.get("/sales-register")
def list_sales_register(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in (UserRole.FINANCE_TEAM, UserRole.FINANCE_CFA_TEAM,
                                  UserRole.CN_TEAM, UserRole.SUPER_ADMIN, UserRole.TENANT_ADMIN):
        raise HTTPException(status_code=403, detail="Access denied")

    q = db.query(SalesRegister).filter(SalesRegister.tenant_id == current_user.tenant_id)
    if search:
        q = q.filter(or_(
            SalesRegister.invoice_no.ilike("%" + search + "%"),
            SalesRegister.dealer_name.ilike("%" + search + "%"),
            SalesRegister.product_name.ilike("%" + search + "%"),
        ))
    total = q.count()
    rows  = q.order_by(SalesRegister.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "records": [{
            "id": str(r.id), "invoice_no": r.invoice_no,
            "invoice_date": r.invoice_date.isoformat() if r.invoice_date else None,
            "dealer_name": r.dealer_name, "dealer_code": r.dealer_code,
            "product_name": r.product_name, "product_code": r.product_code,
            "batch_no": r.batch_no, "quantity": r.quantity,
            "rate": r.rate, "upload_batch": r.upload_batch,
        } for r in rows],
        "total": total, "page": page, "per_page": per_page,
    }


@router.get("/qty-ledger")
def qty_ledger(
    invoice_no: Optional[str] = None,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    q = db.query(QtyUtilizationLedger).filter(QtyUtilizationLedger.tenant_id == current_user.tenant_id)
    if invoice_no:
        q = q.filter(QtyUtilizationLedger.invoice_no == invoice_no)
    rows = q.limit(200).all()
    return [{"invoice_no": r.invoice_no, "product_code": r.product_code, "batch_no": r.batch_no,
             "total_sold_qty": r.total_sold_qty, "claimed_qty": r.claimed_qty, "balance_qty": r.balance_qty} for r in rows]


# ─── Notifications ─────────────────────────────────────────────

@router.get("/notifications")
def get_notifications(
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    notifs = db.query(PDCNNotification).filter(
        PDCNNotification.user_id == current_user.id
    ).order_by(PDCNNotification.created_at.desc()).limit(50).all()
    unread = sum(1 for n in notifs if not n.is_read)
    return {
        "notifications": [{
            "id": str(n.id), "title": n.title, "message": n.message,
            "is_read": n.is_read, "claim_id": str(n.claim_id) if n.claim_id else None,
            "created_at": n.created_at.isoformat() if n.created_at else None,
        } for n in notifs],
        "unread_count": unread,
    }


@router.post("/notifications/{notif_id}/read")
def mark_read(
    notif_id: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    n = db.query(PDCNNotification).filter(
        PDCNNotification.id == notif_id, PDCNNotification.user_id == current_user.id
    ).first()
    if n:
        n.is_read = True
        db.commit()
    return {"message": "Marked as read"}


# ─── Video Upload / Management ─────────────────────────────────

VIDEO_DIR = Path("uploads/videos")
VIDEO_DIR.mkdir(parents=True, exist_ok=True)
MAX_VIDEO_SIZE = 200 * 1024 * 1024  # 200 MB
ALLOWED_VIDEO_MIME = {"video/mp4", "video/webm", "video/quicktime",
                       "video/x-msvideo", "video/mpeg"}


@router.post("/claims/{claim_id}/videos")
async def upload_video(
    claim_id: str,
    request: Request,
    file: UploadFile = File(...),
    description: str = Form(""),
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    from app.models import PDCNVideo

    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id,
        PDCNClaim.tenant_id == current_user.tenant_id,
    ).first()
    if not claim:
        raise HTTPException(404, "Claim not found")

    content = await file.read()
    if len(content) > MAX_VIDEO_SIZE:
        raise HTTPException(400, "Video too large (max 200 MB)")

    ext = Path(file.filename).suffix.lower() or ".mp4"
    safe_name = str(uuid.uuid4()) + ext
    save_path = VIDEO_DIR / safe_name

    with open(save_path, "wb") as f:
        f.write(content)

    video = PDCNVideo(
        claim_id=claim.id,
        tenant_id=current_user.tenant_id,
        filename=safe_name,
        original_name=file.filename,
        file_size=len(content),
        mime_type=file.content_type or "video/mp4",
        description=description,
        uploaded_by=current_user.id,
    )
    db.add(video)
    db.commit()
    db.refresh(video)

    return {
        "id": str(video.id),
        "original_name": video.original_name,
        "filename": video.filename,
        "file_size": video.file_size,
        "description": video.description,
        "created_at": video.created_at.isoformat() if video.created_at else None,
    }


@router.get("/claims/{claim_id}/videos")
def list_videos(
    claim_id: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    from app.models import PDCNVideo
    videos = db.query(PDCNVideo).filter(
        PDCNVideo.claim_id == claim_id,
        PDCNVideo.tenant_id == current_user.tenant_id,
    ).order_by(PDCNVideo.created_at).all()
    return [{
        "id": str(v.id),
        "original_name": v.original_name,
        "filename": v.filename,
        "file_size": v.file_size,
        "description": v.description,
        "mime_type": v.mime_type,
        "created_at": v.created_at.isoformat() if v.created_at else None,
    } for v in videos]


@router.delete("/claims/{claim_id}/videos/{video_id}")
def delete_video(
    claim_id: str,
    video_id: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    from app.models import PDCNVideo
    video = db.query(PDCNVideo).filter(
        PDCNVideo.id == video_id,
        PDCNVideo.claim_id == claim_id,
        PDCNVideo.tenant_id == current_user.tenant_id,
    ).first()
    if not video:
        raise HTTPException(404, "Video not found")
    try:
        (VIDEO_DIR / video.filename).unlink(missing_ok=True)
    except Exception:
        pass
    db.delete(video)
    db.commit()
    return {"message": "Video deleted"}


@router.get("/videos/{filename}")
def stream_video(
    filename: str,
    current_user: User = Depends(get_active_tenant_user),
):
    file_path = VIDEO_DIR / filename
    if not file_path.exists():
        raise HTTPException(404, "Video not found")
    return FileResponse(
        str(file_path),
        media_type="video/mp4",
        headers={"Accept-Ranges": "bytes"},
    )


# ─── Excel Upload for Line Items ──────────────────────────────

@router.post("/claims/{claim_id}/items/excel-upload")
async def upload_items_excel(
    claim_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    """Upload line items from Excel file. Replaces all existing items."""
    import openpyxl, io
    if current_user.role != UserRole.DEALER:
        raise HTTPException(403, "Only dealers can upload line items")

    claim = db.query(PDCNClaim).filter(
        PDCNClaim.id == claim_id,
        PDCNClaim.tenant_id == current_user.tenant_id,
        PDCNClaim.dealer_id == current_user.id,
    ).first()
    if not claim:
        raise HTTPException(404, "Claim not found")
    if claim.status != PDCNStatus.DRAFT:
        raise HTTPException(400, "Can only upload items to a Draft claim")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "Invalid Excel file")

    # Read header row to map columns
    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]

    def col(row, name_variants):
        for name in name_variants:
            for i, h in enumerate(headers):
                if name in h:
                    v = row[i]
                    return str(v).strip() if v is not None else ""
        return ""

    items = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue  # skip empty rows
        invoice_no = col(row, ["zb invoice", "invoice no", "invoice_no"])
        product_name = col(row, ["product name", "product_name", "description"])
        qty_raw = col(row, ["quantity", "qty"])
        invoice_rate = col(row, ["allowed sd", "invoice rate", "invoice_rate", "sd price"])
        claim_rate = col(row, ["selling price", "claim rate", "claim_rate"])
        if not product_name:
            continue
        try:
            qty = int(float(qty_raw)) if qty_raw else 1
        except Exception:
            qty = 1
        diff = ""
        try:
            ir = float(invoice_rate) if invoice_rate else 0
            cr = float(claim_rate) if claim_rate else 0
            diff = str(round(ir - cr, 2))
        except Exception:
            pass
        total = ""
        try:
            total = str(round(qty * float(diff) if diff else 0, 2))
        except Exception:
            pass
        items.append({
            "sap_material_code": col(row, ["sap material", "sap_material", "material code"]),
            "brand_code":        col(row, ["brand code", "brand_code"]),
            "invoice_no":        invoice_no or "MANUAL",
            "is_manual_invoice": col(row, ["manual invoice", "manual"]) in ("yes","true","1","manual"),
            "invoice_date":      col(row, ["zb invoice date", "invoice date", "invoice_date"]),
            "product_name":      product_name,
            "product_code":      col(row, ["product code", "product_code", "sku"]),
            "batch_no":          col(row, ["batch", "batch no", "batch_no"]),
            "quantity":          qty,
            "purchase_price":    col(row, ["purchase price", "purchase_price"]),
            "invoice_rate":      invoice_rate or "0",
            "claim_rate":        claim_rate or "0",
            "difference_amt":    diff,
            "billed_to":         col(row, ["billed to", "billed_to", "hospital"]),
            "billed_date":       col(row, ["billed date", "billed_date"]),
            "dealer_invoice_no": col(row, ["dealer invoice number", "dealer invoice no", "dealer_invoice_no", "dinvoice"]),
            "credit_note_type":  col(row, ["credit note type", "cn type", "cr type"]) or "CR1",
            "total_claim_amt":   total,
        })

    if not items:
        raise HTTPException(400, "No valid rows found in Excel. Check column headers match template.")

    # Delete existing and save new
    db.query(PDCNLineItem).filter(PDCNLineItem.claim_id == claim.id).delete()
    for item in items:
        li = PDCNLineItem(
            claim_id=claim.id,
            sap_material_code=item["sap_material_code"] or None,
            brand_code=item["brand_code"] or None,
            invoice_no=item["invoice_no"],
            is_manual_invoice=item["is_manual_invoice"],
            invoice_date=_parse_date(item["invoice_date"]),
            product_name=item["product_name"],
            product_code=item["product_code"] or None,
            batch_no=item["batch_no"] or None,
            quantity=item["quantity"],
            purchase_price=item["purchase_price"] or None,
            invoice_rate=item["invoice_rate"],
            claim_rate=item["claim_rate"],
            difference_amt=item["difference_amt"],
            billed_to=item["billed_to"] or None,
            billed_date=_parse_date(item["billed_date"]),
            dealer_invoice_no=item["dealer_invoice_no"] or None,
            credit_note_type=item["credit_note_type"],
            total_claim_amt=item["total_claim_amt"],
        )
        db.add(li)

    # Update totals
    claim.total_qty = sum(i["quantity"] for i in items)
    try:
        claim.total_amount = str(round(sum(float(i["total_claim_amt"] or 0) for i in items), 2))
    except Exception:
        pass

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

    return {"uploaded": len(items), "total_qty": claim.total_qty, "total_amount": claim.total_amount}


@router.get("/line-items-template/excel")
def download_line_items_template(
    current_user: User = Depends(get_active_tenant_user),
):
    """Download Excel template for line item upload."""
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Line Items"

    headers = [
        "Material Code", "Brand Code", "Invoice No",
        "If Manual Invoice? (yes/no)", "Invoice Date (DD-MM-YYYY)",
        "Product Name", "Product Code", "Batch No",
        "Quantity", "Purchase Price", "Allowed SD Price",
        "Selling Price (Dealer Invoice)", "Billed To",
        "Billed Date (DD-MM-YYYY)", "Dealer Invoice Number",
        "Credit Note Type (CR1/CR2)"
    ]
    widths = [18,12,20,22,22,28,14,12,10,16,16,24,24,20,20,20]

    hdr_fill = PatternFill("solid", fgColor="1E1B4B")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, i, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36

    # Sample rows
    sample = [
        ["42-5320-064-01","8457","9517004744","no","02-03-2026",
         "PSN TIB STM 5 DEG SZ C L","PSN-001","B001",
         1,27145,23100,27145,"Ponwal Hospital","12/05/2026","188","CR1"],
        ["42-5000-060-01","8445","9517004744","no","02-03-2026",
         "PSN FEM PS CMT CCR NRW SZ 6 L","PSN-002","B001",
         1,28912,25500,28912,"Ponwal Hospital","12/05/2026","188","CR1"],
    ]
    from openpyxl.styles import Border, Side
    bdr = Border(left=Side("thin","CCCCCC"),right=Side("thin","CCCCCC"),
                 top=Side("thin","CCCCCC"),bottom=Side("thin","CCCCCC"))
    alt_fill = PatternFill("solid", fgColor="F0F0FF")
    for ri, row in enumerate(sample, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = bdr
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = alt_fill

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Column","Required","Notes"),
        ("Material Code","No","Internal material code e.g. 42-5320-064-01"),
        ("Brand Code","No","e.g. 8457"),
        ("Invoice No","YES","Zimmer Biomet invoice number"),
        ("If Manual Invoice?","No","yes or no"),
        ("ZB Invoice Date","YES","Format: DD-MM-YYYY"),
        ("Product Name","YES","Full product description"),
        ("Product Code","No","SKU or product code"),
        ("Batch No","No","Batch number"),
        ("Quantity","YES","Integer units"),
        ("Purchase Price","No","Price at which dealer purchased"),
        ("Allowed SD Price","YES","SD / special discount price"),
        ("Selling Price","YES","Price on dealer invoice to hospital"),
        ("Billed To","No","Hospital / customer name"),
        ("Billed Date","No","DD-MM-YYYY"),
        ("Dealer Invoice Number","No","Your invoice number to the hospital"),
        ("Credit Note Type","No","CR1, CR2 etc. Default: CR1"),
    ]
    for r in instructions:
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E1B4B")
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 50

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="LineItems_Template.xlsx"'})


# ─── Serve Attachment (with Auth) ─────────────────────────────

@router.get("/attachments/{filename}")
def serve_attachment(
    filename: str,
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    """Serve a PDCN attachment file — checks the user has access to the claim."""
    # Find the attachment and verify tenant access
    att = db.query(PDCNAttachment).filter(PDCNAttachment.filename == filename).first()
    if not att:
        # Also check credit note files
        cn = db.query(CreditNote).filter(CreditNote.cn_filename == filename).first()
        if cn:
            claim = db.query(PDCNClaim).filter(PDCNClaim.id == cn.claim_id).first()
            if claim and str(claim.tenant_id) != str(current_user.tenant_id):
                raise HTTPException(403, "Access denied")
            fpath = UPLOAD_DIR / filename
            if not fpath.exists():
                raise HTTPException(404, "File not found")
            return FileResponse(str(fpath), filename=cn.cn_original_name or filename)
        raise HTTPException(404, "Attachment not found")

    # Verify tenant
    claim = db.query(PDCNClaim).filter(PDCNClaim.id == att.claim_id).first()
    if not claim or str(claim.tenant_id) != str(current_user.tenant_id):
        raise HTTPException(403, "Access denied")
    if not can_view_claim(claim, current_user):
        raise HTTPException(403, "You do not have access to this claim")

    fpath = UPLOAD_DIR / filename
    if not fpath.exists():
        raise HTTPException(404, "File not found on disk")

    return FileResponse(str(fpath), filename=att.original_name or filename)


@router.get("/sales-register/available-months")
def get_available_months(
    current_user: User = Depends(get_active_tenant_user),
    db: Session = Depends(get_db),
):
    """Return list of months for which sales register has been uploaded.
    Dealers use this to know which months they can create claims for."""
    batches = db.query(SalesRegisterBatch).filter(
        SalesRegisterBatch.tenant_id == current_user.tenant_id,
    ).order_by(SalesRegisterBatch.batch_key.desc()).all()
    return {
        "months": [
            {
                "batch_key":   b.batch_key,
                "month_label": b.month_label or b.batch_key,
                "row_count":   b.row_count,
                "uploaded_at": b.uploaded_at.isoformat() if b.uploaded_at else None,
            }
            for b in batches
        ]
    }
